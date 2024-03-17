from flask import Flask, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
from flask_jwt_extended import create_access_token, jwt_required, get_jwt_identity, JWTManager
from openpyxl import load_workbook

app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['JWT_SECRET_KEY'] = 'my-jwt-secret-key'
app.config['SECRET_KEY'] = 'my-secret-key'
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = False
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
ma = Marshmallow(app)
jwt = JWTManager(app)

class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    age = db.Column(db.Integer, nullable=False)
    city = db.Column(db.String(200), nullable=False)

    def __init__(self, name, age, city):
        self.name = name
        self.age = age
        self.city = city

class ProjectSchema(ma.Schema):
    class Meta:
        fields = ['id', 'name', 'age', 'city']

project_schema = ProjectSchema()
project_schemas = ProjectSchema(many=True)

# Authentication
@app.route('/login', methods=['POST'])
def login():
    username = request.json.get('username', None)
    password = request.json.get('password', None)

    if username == 'user' and password == 'password':
        access_token = create_access_token(identity=username)
        return jsonify(access_token=access_token), 200
    else:
        return jsonify({"msg": "Invalid username or password"}), 401

# Post excel data
@app.route('/post', methods=['POST'])
@jwt_required()
def post_data():
    if request.method == 'POST':

        excel_data = request.files['Mydata']
        Mydata = load_workbook(excel_data)
        Newdata =   Mydata.active

        for i in Newdata.iter_rows(min_row = 2, values_only = True):
            data = Project(name=i[0],age=i[1], city=i[2])
            db.session.add(data)
        db.session.commit()

    return "message: Data retrieve"

# Get all data
@app.route('/get', methods=['GET'])
@jwt_required()
def get_all_data():

    all_posts = Project.query.all()
    result = project_schemas.dump(all_posts)
    return jsonify(result)

# Get single data
@app.route('/get/<int:id>', methods=['GET'])
@jwt_required()
def get_data(id):

    post = Project.query.filter_by(id=id).first()
    result = project_schema.dump(post)
    return jsonify(result)

# Update data
@app.route('/update/<int:id>', methods=['PUT'])
@jwt_required()
def update_data(id):
    post = Project.query.get(id)
    name = request.json['name']
    age = request.json['age']
    city = request.json['city']

    post.name = name
    post.age = age
    post.city = city
    db.session.commit()
    return project_schema.jsonify(post)

# Delte data
@app.route('/delete/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_data(id):
    post = Project.query.get(id)
    db.session.delete(post)
    db.session.commit()
    return project_schema.jsonify(post)

if __name__ == "__main__":
    app.run(debug=True)
